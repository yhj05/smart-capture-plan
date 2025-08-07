#4.엑셀 템플릿 항목에 맞게 작성
from openpyxl import load_workbook
from datetime import datetime
import os
from openpyxl.utils import column_index_from_string

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 💡1. 안전하게 병합하고 값을 쓰는 유틸 함수 정의
def safe_set_and_merge(ws, cell_range, value):
    """
    주어진 셀 범위가 병합되어 있지 않으면 병합 후, 값을 설정합니다.
    이미 병합되어 있으면 값만 설정합니다.
    """
    if cell_range not in [str(range_ref) for range_ref in ws.merged_cells.ranges]:
        ws.merge_cells(cell_range)
    ws[cell_range.split(":")[0]] = value

# 💡2. 실제로 엑셀 작성하는 함수
def write_to_excel(extracted_fields, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb.active

    # 각 필드별로 적절한 셀에 쓰기 (예시)
    if 'Project명' in extracted_fields:
        safe_set_and_merge(ws, "D6:K6", extracted_fields['Project명'])

    if 'Account' in extracted_fields:
        safe_set_and_merge(ws, "D8:F8", extracted_fields['Account'])

    if '고객사 주관 조직' in extracted_fields:
        safe_set_and_merge(ws, "I7:K7", extracted_fields['고객사 주관 조직'])

    if '사업 개요' in extracted_fields:
        safe_set_and_merge(ws, "D10:K10", extracted_fields['사업 개요'])

    if '예상 수행 시작 기간' in extracted_fields:
        safe_set_and_merge(ws, "D11:E11", extracted_fields['예상 수행 시작 기간'])

    if '예상 수행 종료 기간' in extracted_fields:
        safe_set_and_merge(ws, "G11:H11", extracted_fields['예상 수행 종료 기간'])

    if 'RFP 공고' in extracted_fields:
        ws["E22"] = extracted_fields['RFP 공고']

    if '제안서 마감' in extracted_fields:
        ws["G22"] = extracted_fields['제안서 마감']

    if '제안 설명회' in extracted_fields:
        ws["I22"] = extracted_fields['제안 설명회']

    # 저장
    wb.save(output_path)
