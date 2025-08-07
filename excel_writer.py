#4.엑셀 템플릿 항목에 맞게 작성
from openpyxl import load_workbook
from datetime import datetime
import os
from openpyxl.utils import column_index_from_string


#template_path = os.path.join(os.getcwd(), "CapturePlan_Template.xlsx")

def write_to_excel(template_path, output_path, extracted_data):
    wb = load_workbook(template_path, data_only=True, keep_links=False)
    ws = wb["1. 사업 개요"]

    def safe_set_and_merge(cell_range, value):
        # 시작 셀 구하기
        start_cell = cell_range.split(":")[0]
        # 시작 셀에 해당하는 셀 주소 추출
        col_letter = ''.join(filter(str.isalpha, start_cell))
        row_number = int(''.join(filter(str.isdigit, start_cell)))
        

        # 병합 해제
        for merged in list(ws.merged_cells.ranges):
            if start_cell in merged.coord:
                ws.unmerge_cells(str(merged))

        # 병합 해제 후 셀 객체를 새로 가져와 값 할당
        cell = ws.cell(row=row_number, column=column_index_from_string(col_letter))
        cell.value = value    
        
        # 다시 병합
        ws.merge_cells(cell_range)

    # 1. Project명: D~K6
    if extracted_data.get("Project명"):
        safe_set_and_merge("D6:K6", extracted_data["Project명"])

    # 2. Account: D~K7
    if extracted_data.get("Account"):
        safe_set_and_merge("D7:K7", extracted_data["Account"])

    # 2. 사업 개요: D~K10
    if extracted_data.get("사업 개요"):
        safe_set_and_merge("D10:K10", extracted_data["사업 개요"])

    # 3. 고객사 주관 조직: I~K7
    if extracted_data.get("고객사 주관 조직"):
        safe_set_and_merge("I7:K7", extracted_data["고객사 주관 조직"])

    # 4. 수행 기간: 시작 D11:E11, 종료 G11:H11
    if extracted_data.get("예상 수행 시작 기간"):
        safe_set_and_merge("D11:E11", extracted_data["예상 수행 시작 기간"])
    if extracted_data.get("예상 수행 종료 기간"):
        safe_set_and_merge("G11:H11", extracted_data["예상 수행 종료 기간"])

    # 5. 날짜 항목 (병합 필요 없으면 기존대로)
    if extracted_data.get("RFP 공고"):
        ws["E22"] = extracted_data["RFP 공고"]
    if extracted_data.get("제안서 마감"):
        ws["G22"] = extracted_data["제안서 마감"]
    if extracted_data.get("제안 설명회"):
        ws["I22"] = extracted_data["제안 설명회"]

    wb.save(output_path)
