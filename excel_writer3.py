#4.엑셀 템플릿 항목에 맞게 작성
from openpyxl import load_workbook
import openpyxl

def write_to_excel(template_path, output_path, extracted_data):
    
    wb = load_workbook(template_path, data_only=True, keep_links=False)
    ws = wb["1. 사업 개요"]

    def safe_set_and_merge(ws, cell_range, value):
        # 이미 병합된 범위면 병합 해제
        if cell_range in [str(rng) for rng in ws.merged_cells.ranges]:
            ws.unmerge_cells(cell_range)
        # 병합 해제 후, 좌상단 셀을 새로 받아서 값 할당
        min_row, min_col, max_row, max_col = openpyxl.utils.range_boundaries(cell_range)
        cell = ws.cell(row=min_row, column=min_col)
        cell.value = value
        ws.merge_cells(cell_range)

    # 1. Project명: D~K6

    if extracted_data.get("Project명"):
    #     safe_set_and_merge(ws, "D6:K6", extracted_data["Project명"])
        ws["D6"] = extracted_data["Project명"]

    # 2. Account: D~K7
    if extracted_data.get("Account"):
        #safe_set_and_merge(ws, "D7:K7", extracted_data["Account"])
        ws["D7"] = extracted_data["Account"]

    # 2. 사업 개요: D~K10
    if extracted_data.get("사업개요"):
        #safe_set_and_merge(ws, "D10:K10", extracted_data["사업 개요"])
        ws["D10"] = extracted_data["사업개요"]

    # 3. 고객사 주관 조직: I~K7
    if extracted_data.get("고객사 주관 조직"):
        #safe_set_and_merge(ws, "I7:K7", extracted_data["고객사 주관 조직"])
        ws["I7"] = extracted_data["고객사 주관 조직"]

    # 4. 수행 기간: 시작 D11:E11, 종료 G11:H11
    if extracted_data.get("예상 수행 시작 기간"):
        #safe_set_and_merge(ws, "D11:E11", extracted_data["예상 수행 시작 기간"])
        ws["D11"] = extracted_data["예상 수행 시작 기간"]
    if extracted_data.get("예상 수행 종료 기간"):
        #safe_set_and_merge(ws, "G11:H11", extracted_data["예상 수행 종료 기간"])
        ws["G11"] = extracted_data["예상 수행 종료 기간"]

    # 5. 날짜 항목 (병합 필요 없으면 기존대로)
    if extracted_data.get("RFP 공고"):
        ws["E22"] = extracted_data["RFP 공고"]
    if extracted_data.get("제안서 제출 마감일"):
        ws["G22"] = extracted_data["제안서 제출 마감일"]
    if extracted_data.get("제안 설명회"):
        ws["I22"] = extracted_data["제안 설명회"]

    wb.save(output_path)
