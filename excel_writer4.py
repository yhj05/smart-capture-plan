from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

def safe_set_and_merge(ws: Worksheet, cell_range: str, value: str):
    start_cell, end_cell = cell_range.split(":")
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))

    # 병합된 범위 중, 해당 범위와 겹치는 것만 해제
    for merged in list(ws.merged_cells.ranges):
        if start_cell in merged or end_cell in merged:
            try:
                ws.unmerge_cells(str(merged))
            except Exception as e:
                print(f"⚠ 병합 해제 실패: {merged} → {e}")
    
    # 병합 해제 후, 셀에 값 입력
    cell = ws.cell(row=start_row, column=column_index_from_string(start_col))
    cell.value = value

    # 다시 병합
    ws.merge_cells(cell_range)

def write_to_excel(wb: Workbook, output_path: str, extracted_data: dict):
    ws = wb["1. 사업 개요"]

    # 병합 셀 입력
    if extracted_data.get("Project명"):
        safe_set_and_merge(ws, "D6:K6", extracted_data["Project명"])

    if extracted_data.get("Account"):
        safe_set_and_merge(ws, "D7:K7", extracted_data["Account"])

    if extracted_data.get("고객사 주관 조직"):
        safe_set_and_merge(ws, "I7:K7", extracted_data["고객사 주관 조직"])

    if extracted_data.get("사업 개요"):
        safe_set_and_merge(ws, "D10:K10", extracted_data["사업 개요"])

    if extracted_data.get("예상 수행 시작 기간"):
        safe_set_and_merge(ws, "D11:E11", extracted_data["예상 수행 시작 기간"])

    if extracted_data.get("예상 수행 종료 기간"):
        safe_set_and_merge(ws, "G11:H11", extracted_data["예상 수행 종료 기간"])

    # 일반 셀 입력
    if extracted_data.get("RFP 공고"):
        ws["E22"] = extracted_data["RFP 공고"]

    if extracted_data.get("제안서 마감"):
        ws["G22"] = extracted_data["제안서 마감"]

    if extracted_data.get("제안 설명회"):
        ws["I22"] = extracted_data["제안 설명회"]

    # 엑셀 저장
    wb.save(output_path)
